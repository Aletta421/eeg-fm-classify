"""Create deterministic, diagnosis-stratified 80/20 raw-subject splits.

The manifest is generated before any signal preprocessing. A subject has one
row per dataset, so all of that subject's runs/tasks/recordings stay together.
"""

import argparse
import csv
import hashlib
import random
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from scipy.io import loadmat

ROOT = Path(__file__).resolve().parent.parent

TDBRAIN_LABELS = {
    "HEALTHY": (0, "control"),
    "MDD": (1, "depression"),
    "ADHD": (2, "adhd"),
    "ADD": (2, "adhd"),
}


def add_record(records, dataset, subject_id, label, diagnosis_type, n_records=1,
               payload_ready=True):
    key = (dataset, str(subject_id))
    value = records.get(key)
    if value:
        if (value["label"], value["diagnosis_type"]) != (label, diagnosis_type):
            raise ValueError(f"conflicting labels for {dataset}/{subject_id}")
        value["n_records"] += n_records
        value["payload_ready"] = value["payload_ready"] and payload_ready
        return
    records[key] = {
        "dataset": dataset,
        "subject_id": str(subject_id),
        "label": int(label),
        "diagnosis_type": diagnosis_type,
        "n_records": int(n_records),
        "payload_ready": bool(payload_ready),
    }


def read_modma(records):
    root = ROOT / "data" / "MODMA"
    label_map = {}
    for workbook_path in root.rglob("subjects_information_*.xlsx"):
        sheet = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True).active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            subject_id = str(row[0]).strip().zfill(8)
            group = str(row[1]).strip().upper()
            if group == "MDD":
                label_map[subject_id] = (1, "depression")
            elif group == "HC":
                label_map[subject_id] = (0, "control")

    counts = Counter()
    for pattern in ("*.mat", "*.raw", "*.txt"):
        for path in root.rglob(pattern):
            subject_id = path.stem[:8]
            if subject_id in label_map:
                counts[subject_id] += 1
    for subject_id, n_records in counts.items():
        add_record(records, "MODMA", subject_id, *label_map[subject_id], n_records)


def read_ieee(records):
    root = ROOT / "data" / "IEEE_ADHD"
    for path in root.rglob("*.mat"):
        diagnosis = "adhd" if "adhd" in path.parent.name.lower() else "control"
        add_record(records, "IEEE_ADHD", path.stem,
                   1 if diagnosis == "adhd" else 0, diagnosis)


def read_mendeley(records):
    root = ROOT / "data" / "Mendeley_ADHD"
    groups = {
        "FC": (0, "control"), "MC": (0, "control"),
        "FADHD": (1, "adhd"), "MADHD": (1, "adhd"),
    }
    for path in root.rglob("*.mat"):
        group = path.stem
        if group not in groups:
            continue
        mat = loadmat(path)
        key = next(key for key in mat if not key.startswith("__"))
        cells = mat[key]
        n_subjects = cells[0, 0].shape[0]
        for index in range(n_subjects):
            if group == "FADHD" and index == 6:
                continue  # documented all-zero corrupted subject
            subject_id = f"{group}_{index + 1:02d}"
            n_records = cells.shape[1]
            add_record(records, "Mendeley", subject_id, *groups[group], n_records)


def is_annex_pointer(path):
    if path.stat().st_size > 1024:
        return False
    try:
        return ".git/annex/objects" in path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return True


def read_openneuro(records):
    root = ROOT / "data" / "OpenNeuro_ds003478"
    labels = {}
    with (root / "participants.tsv").open("r", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            try:
                bdi = int(row["BDI"])
            except (KeyError, TypeError, ValueError):
                continue
            if bdi >= 14:
                labels[row["participant_id"]] = (1, "depression")
            elif bdi <= 7:
                labels[row["participant_id"]] = (0, "control")

    for path in root.rglob("*_eeg.set"):
        subject_id = next((part for part in path.parts
                           if part.startswith("sub-") and "_" not in part), None)
        if subject_id in labels:
            fdt = path.with_suffix(".fdt")
            ready = (not is_annex_pointer(path) and fdt.exists()
                     and not is_annex_pointer(fdt))
            add_record(records, "OpenNeuro_ds003478", subject_id,
                       *labels[subject_id], 1, payload_ready=ready)


def read_tdbrain(records):
    data_root = ROOT / "data" / "TDBRAIN" / "TDBRAIN_Dataset_V3_1"
    workbook = ROOT / "data" / "TDBRAIN" / "TDBRAIN_participants_V3.xlsx"
    sheet = openpyxl.load_workbook(workbook, read_only=True, data_only=True).active
    headers = [cell.value for cell in sheet[1]]
    id_col = headers.index("TDBRAIN_ID")
    indication_col = headers.index("indication")
    status_col = headers.index("formal_status")
    labels = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[id_col] is None:
            continue
        indication = str(row[indication_col]).strip() if row[indication_col] is not None else ""
        status = str(row[status_col]).strip() if row[status_col] is not None else ""
        source = indication if indication.upper() not in {"", "NAN", "UNKNOWN"} else status
        primary = source.split("/")[0].strip().upper()
        if primary in TDBRAIN_LABELS:
            labels[str(row[id_col]).strip()] = TDBRAIN_LABELS[primary]

    counts = Counter(path.name.split("_")[0] for path in data_root.rglob("*.bdf"))
    for subject_id, n_records in counts.items():
        if subject_id in labels:
            add_record(records, "TDBRAIN", subject_id, *labels[subject_id], n_records)


def assign_splits(records, seed, test_fraction):
    strata = defaultdict(list)
    for record in records.values():
        strata[(record["dataset"], record["diagnosis_type"])].append(record)

    for (dataset, diagnosis), group in sorted(strata.items()):
        group.sort(key=lambda item: item["subject_id"])
        digest = hashlib.sha256(f"{seed}:{dataset}:{diagnosis}".encode()).digest()
        rng = random.Random(int.from_bytes(digest[:8], "big"))
        rng.shuffle(group)
        if len(group) < 2:
            n_test = 0
        else:
            n_test = max(1, min(len(group) - 1, int(len(group) * test_fraction + 0.5)))
        for index, record in enumerate(group):
            record["split"] = "test" if index < n_test else "train"
            record["seed"] = seed


def validate(records):
    seen = {}
    for record in records.values():
        key = (record["dataset"], record["subject_id"])
        if key in seen and seen[key] != record["split"]:
            raise ValueError(f"subject leakage detected: {key}")
        seen[key] = record["split"]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", default=str(ROOT / "data" / "splits.csv"))
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--test-fraction", type=float, default=0.2)
    args = parser.parse_args()
    if not 0 < args.test_fraction < 1:
        parser.error("--test-fraction must be between 0 and 1")

    records = {}
    for reader in (read_modma, read_ieee, read_mendeley, read_openneuro, read_tdbrain):
        reader(records)
    assign_splits(records, args.seed, args.test_fraction)
    validate(records)

    columns = ["dataset", "subject_id", "split", "label", "diagnosis_type",
               "n_records", "payload_ready", "seed"]
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(sorted(records.values(), key=lambda r: (r["dataset"], r["subject_id"])))

    print(f"Wrote {len(records)} raw-subject assignments to {output}")
    summary = Counter((r["dataset"], r["split"], r["diagnosis_type"])
                      for r in records.values())
    for key, count in sorted(summary.items()):
        print(f"  {key[0]:22s} {key[1]:5s} {key[2]:10s}: {count}")
    missing = sum(not r["payload_ready"] for r in records.values())
    if missing:
        print(f"  WARNING: {missing} subjects have metadata only (payload unavailable)")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
