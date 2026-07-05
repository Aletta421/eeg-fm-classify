"""
TDBRAIN Dataset V3 加载器

数据来源: https://brainclinics.com/resources/
完整数据集: TDBRAIN_Dataset_V3_1 — 1300 受试者 (DISCOVERY set)
  - 每受试者最多 2 种条件: restEC (静息闭眼) + oddball (听觉 oddball)
  - 格式: BioSemi .bdf, 采样率 ~500 Hz
  - 标签来自 TDBRAIN_participants_V3.xlsx，按 sub-XXXXXXXX ID 精确匹配

标签策略:
  - 优先使用 indication 列（纯诊断如 MDD/ADHD/HEALTHY）
  - indication 为 nan 时回退 formal_status
  - 复合诊断（如 MDD/ADHD）取第一个作为主标签
  - 二分类: MDD→1, HEALTHY→0; 同时也保留 diagnosis_type 用于多类别

使用方式:
    python load_tdbrain.py --data_dir ../data/TDBRAIN/TDBRAIN_Dataset_V3_1 \
                           --output_dir ../data/processed/TDBRAIN \
                           --xlsx ../data/TDBRAIN/TDBRAIN_participants_V3.xlsx
"""

import argparse
import sys
from pathlib import Path
from typing import Tuple, Dict, Any, List, Optional
import warnings

import numpy as np
import openpyxl

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from base_loader import BaseEEGLoader

# 非 EEG 通道（需过滤）
NON_EEG = {
    "Status", "STATUS", "status", "Trigger", "TRIGGER",
    "EOG", "ECG", "EMG", "EKG",
    "Erbs", "Mass", "VPVA", "VNVB", "HPHL", "HNHR",
}

# indication / formal_status → (label_id, diagnosis_type) 映射
# label_id: 0=healthy, 1=depression, 2=adhd, 3+=others
INDICATION_MAP = {
    "HEALTHY":       (0, "control"),
    "MDD":           (1, "depression"),
    "ADHD":          (2, "adhd"),
    "ADHD ":         (2, "adhd"),  # trailing space variant
    "ADD":           (2, "adhd"),
    "OCD":           (3, "ocd"),
    "INSOMNIA":      (4, "insomnia"),
    "TINNITUS":      (5, "tinnitus"),
    "Tinnitus":      (5, "tinnitus"),
    "PARKINSON":     (6, "parkinson"),
    "CHRONIC PAIN":  (7, "chronic_pain"),
    "BURNOUT":       (8, "burnout"),
    "Dyslexia":      (9, "dyslexia"),
    "BIPOLAR":       (10, "bipolar"),
    "ANXIETY":       (11, "anxiety"),
    "SMC":           (12, "smc"),
    "DEPERSONALIZATION": (13, "depersonalization"),
    "WHIPLASH":      (14, "whiplash"),
    "MIGRAINE":      (15, "migraine"),
    "TBI":           (16, "tbi"),
    "ASPERGER":      (17, "asperger"),
    "ASD":           (18, "asd"),
    "PANIC":         (19, "panic"),
    "PTSD":          (20, "ptsd"),
    "EPILEPSY":      (21, "epilepsy"),
    "STROKE":        (22, "stroke"),
    "PDD NOS":       (23, "pdd_nos"),
    "CONVERSION DX": (24, "conversion_dx"),
}


class TDBrainLoader(BaseEEGLoader):
    """TDBRAIN Dataset V3 加载器。

    读取 .bdf，按 sub-XXXXXXXX ID 匹配 xlsx 标签。
    """

    dataset_name = "TDBRAIN"

    def __init__(self, config_path: str = "../configs/preprocess_config.yaml",
                 xlsx_path: Optional[str] = None):
        super().__init__(config_path)
        self._labels: Dict[str, Tuple[int, str]] = {}
        if xlsx_path:
            self._load_xlsx(xlsx_path)

    # ================================================================
    # 子类实现
    # ================================================================

    def _read_file(self, file_path: str) -> Tuple[np.ndarray, Dict[str, Any]]:
        """读取 BioSemi .bdf 文件。"""
        import mne

        file_path = Path(file_path)
        fname = file_path.name  # e.g. sub-87957837_ses-1_task-restEC_eeg.bdf

        # 解析 subject_id 和 task
        parts = fname.replace(".bdf", "").split("_")
        subject_id = parts[0]
        task = "unknown"
        for p in parts:
            if p.startswith("task-"):
                task = p[5:]
                break

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            raw = mne.io.read_raw_bdf(str(file_path), preload=False, verbose=False)
            raw.load_data()

        all_data = raw.get_data()
        fs = raw.info["sfreq"]
        all_ch_names = raw.info["ch_names"]

        # 过滤非 EEG 通道
        eeg_indices = [
            i for i, ch in enumerate(all_ch_names) if ch not in NON_EEG
        ]
        data = all_data[eeg_indices]
        ch_names = [all_ch_names[i] for i in eeg_indices]

        return data, {
            "subject_id": subject_id,
            "fs": fs,
            "ch_names": ch_names,
            "task": task,
            "format": "bdf",
        }

    def _parse_label(self, meta: Dict[str, Any]) -> Tuple[int, str]:
        """从 xlsx 标签缓存中查找。"""
        subject_id = meta["subject_id"]
        return self._labels.get(subject_id, (-1, "unknown"))

    def _extract_channels(
        self, data: np.ndarray, meta: Dict[str, Any]
    ) -> Tuple[List[str], np.ndarray]:
        """提取通道名并匹配标准 10-20 蒙太奇坐标。"""
        import mne

        ch_names = meta.get("ch_names", [f"Ch{i+1}" for i in range(data.shape[0])])

        try:
            montage = mne.channels.make_standard_montage("standard_1020")
            pos = montage.get_positions()["ch_pos"]
            positions = np.array([
                pos.get(ch, [0.0, 0.0, 0.0]) for ch in ch_names
            ])
        except Exception:
            positions = np.zeros((data.shape[0], 3))

        return ch_names, positions

    # ================================================================
    # 标签加载
    # ================================================================

    def _load_xlsx(self, xlsx_path: str):
        """从 xlsx 加载标签，按 sub-XXXXXXXX ID 精确匹配。

        优先 indication 列，nan 时回退 formal_status。
        复合诊断取第一个。
        """
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        id_col = headers.index("TDBRAIN_ID")
        ind_col = headers.index("indication")
        status_col = headers.index("formal_status")

        loaded = 0
        unknown = 0
        composite = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            subj_id = str(row[id_col]) if row[id_col] is not None else None
            if subj_id is None:
                continue

            indication = str(row[ind_col]) if row[ind_col] is not None else "nan"
            formal_status = str(row[status_col]) if row[status_col] is not None else "nan"

            # 优先 indication，回退 formal_status
            label_source = indication if indication != "nan" else formal_status
            if label_source == "nan" or label_source == "UNKNOWN":
                unknown += 1
                continue

            # 处理复合诊断 (如 MDD/ADHD) — 取第一个
            if "/" in label_source:
                composite += 1
            primary = label_source.split("/")[0].strip()

            if primary in INDICATION_MAP:
                label, diag_type = INDICATION_MAP[primary]
                self._labels[subj_id] = (label, diag_type)
                loaded += 1
            else:
                unknown += 1

        print(f"  [TDBRAIN] 标签: {loaded} 已加载, {composite} 复合诊断, "
              f"{unknown} 未知/跳过")


# ================================================================
# 命令行入口
# ================================================================

def main():
    parser = argparse.ArgumentParser(description="TDBRAIN Dataset V3 预处理")
    parser.add_argument("--data_dir", type=str,
                        default="../data/TDBRAIN/TDBRAIN_Dataset_V3_1",
                        help="TDBRAIN 原始数据目录")
    parser.add_argument("--output_dir", type=str,
                        default="../data/processed/TDBRAIN",
                        help="预处理后输出目录")
    parser.add_argument("--xlsx", type=str,
                        default="../data/TDBRAIN/TDBRAIN_participants_V3.xlsx",
                        help="受试者标签 xlsx 路径")
    parser.add_argument("--config", type=str,
                        default="../configs/preprocess_config.yaml",
                        help="配置文件路径")
    parser.add_argument("--conditions", type=str, default="restEC,oddball",
                        help="处理的条件，逗号分隔 (默认: restEC,oddball)")
    parser.add_argument("--limit", type=int, default=0,
                        help="限制受试者数量 (0=全部)")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir)
    conditions = [c.strip() for c in args.conditions.split(",")]

    # 扫描所有 .bdf 文件
    bdf_files = sorted(data_dir.glob("sub-*/*/eeg/*.bdf"))
    if not bdf_files:
        print(f"❌ 未找到 .bdf 文件: {data_dir}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"TDBRAIN Dataset V3 预处理")
    print(f"{'='*60}")

    loader = TDBrainLoader(args.config, xlsx_path=args.xlsx)

    # 按 (subject_id, task) 组织
    from collections import defaultdict
    subjects: Dict[str, Dict[str, Path]] = defaultdict(dict)

    for bf in bdf_files:
        fname = bf.name
        sid = fname.split("_")[0]
        for part in fname.split("_"):
            if part.startswith("task-"):
                task = part[5:]
                if task in conditions:
                    subjects[sid][task] = bf

    print(f"  受试者: {len(subjects)}")
    print(f"  .bdf 文件: {len(bdf_files)}")
    print(f"  条件: {conditions}\n")

    total_files = 0
    skipped = 0
    excluded = 0

    subject_ids = sorted(subjects.keys())
    if args.limit > 0:
        subject_ids = subject_ids[:args.limit]
        print(f"  ⚠ 限制处理 {args.limit} 个受试者\n")

    for i, sid in enumerate(subject_ids):
        label, diag_type = loader._parse_label({"subject_id": sid})
        if label == -1:
            excluded += 1
            continue

        for task, bdf_path in subjects[sid].items():
            try:
                result = loader.process(str(bdf_path))
                out_dir = output_dir / sid / task
                loader.save(result, str(out_dir))
                total_files += 1
            except Exception as e:
                print(f"  ✗ {sid}/{task}: {e}", file=sys.stderr)
                skipped += 1

        if (i + 1) % 100 == 0:
            print(f"  已处理 {i+1}/{len(subject_ids)} ({total_files} 文件)...")

    print(f"\n✓ TDBRAIN 预处理完成: "
          f"{len(subject_ids) - excluded} 受试者, {total_files} 文件")
    print(f"  排除 (无标签): {excluded}")
    print(f"  跳过 (错误): {skipped}")
    print(f"  输出目录: {output_dir}")


if __name__ == "__main__":
    main()
